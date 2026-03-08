import tkinter as tk
from tkinter import messagebox, filedialog
import json
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

# ================= PATH CONFIG =================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

USERS_DB = os.path.join(BASE_DIR, "users.json")
DATA_FILE = os.path.join(BASE_DIR, "water_data.json")
EXCEL_PATH_FILE = os.path.join(BASE_DIR, "excel_path.json")
# ===============================================

# ================= THEMES ======================
LIGHT_THEME = {
    "bg": "#F2F2F7",
    "card": "#FFFFFF",
    "text": "#1C1C1E",
    "sub": "#6E6E73",
    "primary": "#0A84FF",
    "success": "#34C759",
    "danger": "#FF3B30",
    "track": "#E5E5EA"
}

DARK_THEME = {
    "bg": "#1C1C1E",
    "card": "#2C2C2E",
    "text": "#FFFFFF",
    "sub": "#8E8E93",
    "primary": "#0A84FF",
    "success": "#30D158",
    "danger": "#FF453A",
    "track": "#3A3A3C"
}
# ===============================================

FONT_TITLE = ("Segoe UI", 18, "bold")
FONT_BIG = ("Segoe UI", 15, "bold")
FONT_MAIN = ("Segoe UI", 13)


class WaterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Hydration Tracker")
        self.root.geometry("430x760")
        self.root.resizable(False, False)

        self.theme = LIGHT_THEME
        self.current_user = None
        self.excel_path = self.load_excel_path()

        self.show_login()

    # ================= AUTH ======================
    def show_login(self):
        self.clear()
        self.root.configure(bg=self.theme["bg"])

        card = self.card()
        tk.Label(
            card,
            text="Hydration Tracker",
            font=FONT_TITLE,
            bg=self.theme["card"],
            fg=self.theme["text"]
        ).pack(pady=10)

        self.login_var = tk.StringVar()
        self.pass_var = tk.StringVar()

        self.input(card, "Username", self.login_var)
        self.input(card, "Password", self.pass_var, show="*")

        self.button(card, "Login", self.theme["primary"], self.login).pack(pady=6)
        self.button(card, "Register", self.theme["success"], self.show_register).pack()

    def show_register(self):
        self.clear()
        card = self.card()

        tk.Label(
            card,
            text="Create Account",
            font=FONT_TITLE,
            bg=self.theme["card"],
            fg=self.theme["text"]
        ).pack(pady=10)

        self.reg_login = tk.StringVar()
        self.reg_pass = tk.StringVar()

        self.input(card, "Username", self.reg_login)
        self.input(card, "Password", self.reg_pass, show="*")

        self.button(card, "Create Account", self.theme["success"], self.register).pack(pady=6)
        self.button(card, "Back", self.theme["primary"], self.show_login).pack()

    def login(self):
        users = self.load_users()
        if users.get(self.login_var.get()) == self.pass_var.get():
            self.current_user = self.login_var.get()
            self.load_data()
            self.show_main()
        else:
            messagebox.showerror("Error", "Invalid username or password")

    def register(self):
        users = self.load_users()
        if self.reg_login.get() in users:
            messagebox.showerror("Error", "Account already exists")
            return

        users[self.reg_login.get()] = self.reg_pass.get()
        self.save_users(users)

        messagebox.showinfo("Success", "Account created successfully")
        self.show_login()

    def load_users(self):
        if not os.path.exists(USERS_DB):
            return {}
        with open(USERS_DB, "r", encoding="utf-8") as f:
            return json.load(f)

    def save_users(self, users):
        with open(USERS_DB, "w", encoding="utf-8") as f:
            json.dump(users, f, indent=4, ensure_ascii=False)

    # ================= MAIN UI ===================
    def show_main(self):
        self.clear()
        self.root.configure(bg=self.theme["bg"])

        header = tk.Frame(self.root, bg=self.theme["bg"])
        header.pack(fill="x", pady=10)

        tk.Label(
            header,
            text="💧 Hydration Tracker",
            font=FONT_TITLE,
            bg=self.theme["bg"],
            fg=self.theme["text"]
        ).pack(side="left", padx=16)

        self.button(
            header,
            "🌙 / ☀️",
            self.theme["primary"],
            self.toggle_theme
        ).pack(side="right", padx=16)

        card = self.card()
        self.weight_var = tk.StringVar(value=str(self.data["weight"]))
        self.input(card, "Weight (kg)", self.weight_var)
        self.button(card, "Calculate daily goal", self.theme["primary"], self.calc_goal).pack()

        card2 = self.card()
        self.stat = tk.Label(
            card2,
            font=FONT_BIG,
            bg=self.theme["card"],
            fg=self.theme["text"]
        )
        self.stat.pack(pady=6)

        self.canvas = tk.Canvas(
            card2,
            width=300,
            height=18,
            bg=self.theme["track"],
            highlightthickness=0
        )
        self.canvas.pack(pady=6)

        self.bar = self.canvas.create_rectangle(0, 0, 0, 18, fill="#FF3B30", width=0)

        card3 = self.card()
        self.button(card3, "+250 ml", self.theme["success"], lambda: self.add_water(250)).pack(side="left", padx=6)
        self.button(card3, "+500 ml", self.theme["success"], lambda: self.add_water(500)).pack(side="left", padx=6)

        self.button(
            self.root,
            "Finish Day",
            self.theme["danger"],
            self.finish_day
        ).pack(pady=20)

        self.refresh()

    # ================= LOGIC =====================
    def toggle_theme(self):
        self.theme = DARK_THEME if self.theme == LIGHT_THEME else LIGHT_THEME
        self.show_main()

    def load_data(self):
        if not os.path.exists(DATA_FILE):
            self.data = {
                "weight": 70,
                "goal": 2100,
                "history": []
            }
            self.save_data()
        else:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                self.data = json.load(f)

    def save_data(self):
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=4, ensure_ascii=False)

    def calc_goal(self):
        self.data["weight"] = float(self.weight_var.get())
        self.data["goal"] = int(self.data["weight"] * 30)
        self.save_data()
        self.refresh()

    def add_water(self, amount):
        self.data["history"].append({
            "user": self.current_user,
            "date": datetime.now().strftime("%Y-%m-%d"),
            "time": datetime.now().strftime("%H:%M:%S"),
            "amount": amount
        })
        self.save_data()
        self.refresh()

    def refresh(self):
        current = sum(x["amount"] for x in self.data["history"])
        goal = self.data["goal"]
        percent = min(current / goal, 1)

        color = "#FF3B30" if percent < 0.4 else "#FFD60A" if percent < 0.8 else "#34C759"

        self.stat.config(text=f"{current} / {goal} ml")
        self.canvas.itemconfig(self.bar, fill=color)
        self.canvas.coords(self.bar, 0, 0, int(300 * percent), 18)

    # ================= EXCEL =====================
    def finish_day(self):
        if not self.data["history"]:
            return

        if messagebox.askyesno("Excel Export", "Save today's data to Excel?"):
            self.save_excel()

        self.data["history"] = []
        self.save_data()
        self.refresh()

    def save_excel(self):
        if not self.excel_path or not os.path.exists(self.excel_path):
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Excel File"
            )
            if not path:
                return

            self.excel_path = path
            with open(EXCEL_PATH_FILE, "w", encoding="utf-8") as f:
                json.dump({"path": path}, f, indent=4, ensure_ascii=False)

        wb = load_workbook(self.excel_path) if os.path.exists(self.excel_path) else Workbook()
        ws = wb.active
        ws.title = "Water History"

        if ws.max_row == 1:
            ws.append(["User", "Date", "Time", "Amount (ml)"])
            for i in range(1, 5):
                ws.column_dimensions[get_column_letter(i)].width = 22

        font = Font(name="Times New Roman", size=14)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for entry in self.data["history"]:
            ws.append([
                entry["user"],
                entry["date"],
                entry["time"],
                entry["amount"]
            ])

        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.border = border

        wb.save(self.excel_path)

    def load_excel_path(self):
        if os.path.exists(EXCEL_PATH_FILE):
            with open(EXCEL_PATH_FILE, "r", encoding="utf-8") as f:
                return json.load(f).get("path")
        return None

    # ================= UI HELPERS =================
    def clear(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def card(self):
        frame = tk.Frame(
            self.root,
            bg=self.theme["card"],
            highlightbackground="#3A3A3C",
            highlightthickness=1
        )
        frame.pack(padx=16, pady=8, fill="x")
        return frame

    def input(self, parent, label, var, show=None):
        tk.Label(
            parent,
            text=label,
            bg=self.theme["card"],
            fg=self.theme["sub"]
        ).pack()

        tk.Entry(
            parent,
            textvariable=var,
            show=show,
            font=FONT_BIG,
            justify="center",
            relief="flat",
            bg=self.theme["track"],
            fg=self.theme["text"]
        ).pack(pady=6, ipady=6)

    def button(self, parent, text, color, command):
        return tk.Button(
            parent,
            text=text,
            bg=color,
            fg="white",
            font=FONT_MAIN,
            relief="flat",
            command=command
        )


# ================= RUN APP ======================
if __name__ == "__main__":
    root = tk.Tk()
    WaterApp(root)
    root.mainloop()
